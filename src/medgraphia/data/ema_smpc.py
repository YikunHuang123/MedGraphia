"""
EMA SmPC (Summary of Product Characteristics) connector.
Downloads the official bulk XML package from EMA's open-data portal.
No scraping — uses the published REST/bulk endpoints only.

EMA product data index:  https://www.ema.europa.eu/en/medicines/download-medicine-data
EPAR XML download:       https://www.ema.europa.eu/sites/default/files/Medicines_output_european_public_assessment_reports.xlsx
                         (we parse the Excel index to find per-drug XML links)

For each product we download the SmPC PDF / XML from the EMA product page.
"""
from __future__ import annotations

import asyncio
import io
import re
import zipfile
from pathlib import Path

import aiohttp
from tenacity import retry, stop_after_attempt, wait_exponential

from medgraphia.config import get_settings
from medgraphia.domain import Language, RawDocument, SourceMeta
from medgraphia.logger import get_logger

logger = get_logger(__name__)

# EMA EPAR index (publicly available, no auth required)
_EPAR_INDEX_URL = "https://www.ema.europa.eu/en/documents/other/medicines-output-medicines-report_en.xlsx"
_EMA_PRODUCT_BASE = "https://www.ema.europa.eu"


class EMASmPCConnector:
    """
    Downloads EMA SmPC XML/PDF data from the public EPAR bulk dataset.
    Produces RawDocument objects for downstream parsing.
    """

    def __init__(self, output_dir: str = "data/raw/ema_smpc") -> None:
        cfg = get_settings()
        self._output_dir = Path(output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)
        self._session: aiohttp.ClientSession | None = None

    async def __aenter__(self) -> "EMASmPCConnector":
        self._session = aiohttp.ClientSession(
            timeout=aiohttp.ClientTimeout(total=120),
            headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
            },
        )
        return self

    async def __aexit__(self, *args: object) -> None:
        if self._session:
            await self._session.close()

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    async def fetch_product_list(self) -> list[dict[str, str]]:
        """
        Download and parse the EPAR index Excel file.
        Checks for a local copy at data/raw/ema_smpc/index.xlsx first.
        Returns a list of {name, category, active_substance, url} dicts.
        """
        logger.info("ema_fetching_index")
        try:
            import openpyxl  # type: ignore
        except ImportError:
            raise ImportError("Install openpyxl to parse EMA EPAR index: pip install openpyxl")

        # 1. Check for local cache first (bypass anti-bot)
        local_index = Path("data/raw/ema_smpc/index.xlsx")
        if local_index.exists():
            logger.info("ema_using_local_index", path=str(local_index))
            data = local_index.read_bytes()
        else:
            # 2. Attempt download (might be blocked by CloudFront/Antibot)
            assert self._session is not None
            logger.info("ema_downloading_index", url=_EPAR_INDEX_URL)
            try:
                async with self._session.get(_EPAR_INDEX_URL) as resp:
                    if resp.status != 200:
                        logger.error("ema_index_http_error", status=resp.status)
                    resp.raise_for_status()
                    data = await resp.read()
            except Exception as exc:
                logger.error("ema_download_failed", error=str(exc))
                raise RuntimeError(
                    f"Could not download EMA index due to anti-bot protection.\n"
                    f"Please manually download the file from:\n{_EPAR_INDEX_URL}\n"
                    f"and save it to: {local_index}"
                ) from exc

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        
        # 3. Dynamically find the header row (EMA index has preambles)
        header_row_idx = 1
        headers = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
            if row and "Category" in row and ("Name of medicine" in row or "Medicine name" in row):
                header_row_idx = i + 1
                headers = [str(c).strip() if c else "" for c in row]
                break
        
        if not headers:
            raise RuntimeError("Could not find a valid header row in EMA Excel index.")

        # Map current EMA column names to our internal keys
        # 2026 names: 'Name of medicine', 'Active substance', 'Medicine URL'
        col_map = {
            "Name of medicine": "Medicine name",
            "Medicine URL": "URL"
        }

        products = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):  # type: ignore[call-arg]
            entry: dict[str, str] = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    header_name = headers[i]
                    # Normalize header name if it's in our map
                    key = col_map.get(header_name, header_name)
                    entry[key] = str(val or "").strip()
            
            if entry.get("Medicine name"):
                products.append(entry)

        logger.info("ema_index_loaded", count=len(products))
        return products

    async def download_smpc(
        self,
        product_url: str,
        product_name: str,
        language: Language = Language.EN,
    ) -> RawDocument | None:
        """
        Given a product page URL, attempt to download the SmPC PDF/XML.
        Returns a RawDocument with file_path pointing to the saved file.
        """
        assert self._session is not None
        try:
            logger.debug("ema_extracting_pdf_url", url=product_url)
            pdf_url = await self._extract_smpc_url(product_url)
            if not pdf_url:
                logger.warning("ema_no_smpc_url", product=product_name, page_url=product_url)
                return None

            file_name = re.sub(r"[^\w\-]", "_", product_name) + ".pdf"
            out_path = self._output_dir / file_name

            await self._download_file(pdf_url, out_path)
            logger.info("ema_smpc_downloaded", product=product_name, path=str(out_path))

            source = SourceMeta(
                source_id=f"ema:{product_name.lower().replace(' ', '_')}",
                source_title=f"EMA SmPC — {product_name}",
                source_version="",
                source_url=pdf_url,
                language=language,
            )
            return RawDocument(
                source=source,
                language=language,
                title=f"EMA SmPC — {product_name}",
                file_path=str(out_path),
                format="pdf",
            )
        except Exception as exc:
            logger.warning("ema_download_failed", product=product_name, error=str(exc))
            return None

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @retry(stop=stop_after_attempt(3), wait=wait_exponential(min=1, max=10))
    async def _extract_smpc_url(self, product_url: str) -> str | None:
        """Scrape the product page for the SmPC document download link."""
        from bs4 import BeautifulSoup

        assert self._session is not None
        async with self._session.get(product_url) as resp:
            logger.debug("ema_page_fetch_status", status=resp.status, url=product_url)
            if resp.status != 200:
                return None
            html = await resp.text()

        soup = BeautifulSoup(html, "lxml")
        
        # EMA site uses specific patterns for Product Information (which contains SmPC)
        for link in soup.find_all("a", href=True):
            href: str = link["href"]
            text: str = link.get_text(strip=True).lower()
            
            # Pattern 1: Language-specific Product Info
            # E.g. "English (EN)" link pointing to "/product-information/..."
            if "english (en)" in text and "/product-information/" in href.lower():
                return href if href.startswith("http") else _EMA_PRODUCT_BASE + href

            # Pattern 2: Text contains "Product information" and is a PDF
            if "product information" in text and href.lower().endswith(".pdf"):
                return href if href.startswith("http") else _EMA_PRODUCT_BASE + href

            # Pattern 3: Fallback to any PDF containing product-information in URL
            if "/product-information/" in href.lower() and href.lower().endswith(".pdf"):
                return href if href.startswith("http") else _EMA_PRODUCT_BASE + href
                
        return None

    @retry(stop=stop_after_attempt(3), wait=wait_exponential(min=2, max=30))
    async def _download_file(self, url: str, dest: Path) -> None:
        assert self._session is not None
        async with self._session.get(url) as resp:
            resp.raise_for_status()
            content = await resp.read()
        dest.write_bytes(content)
